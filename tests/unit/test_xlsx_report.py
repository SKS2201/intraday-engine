from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from src.engine.stage1_openingrange import Stage1RunReport
from src.reports.xlsx_report import write_stage1_report_xlsx


def test_write_stage1_report_xlsx_creates_expected_sheets(tmp_path: Path):
    report = Stage1RunReport(
        message_text="x",
        run_summary={"mode": "TEST", "decision": "NO TRADE"},
        long_rows=[{"rank": "1", "symbol": "INFY"}],
        short_rows=[{"rank": "1", "symbol": "TCS"}],
        metrics_rows=[{"symbol": "INFY", "vwap": "100.00"}],
        validation_rows=[{"issue": "none"}],
        process_log=["step_a", "step_b"],
        workbook_name="stage1_test.xlsx",
    )
    path = write_stage1_report_xlsx(report, str(tmp_path), datetime.fromisoformat("2026-02-06T09:30:00+05:30"))
    assert Path(path).exists()
    wb = load_workbook(path)
    assert set(wb.sheetnames) == {
        "Run_Summary",
        "Long_Ranked",
        "Short_Ranked",
        "Metrics",
        "Validation",
        "Process_Log",
    }
